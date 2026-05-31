# -*- coding: utf-8 -*-
"""
运行 ems_solver_v3 并导出结果到 result2.xlsx
"""

import sys
sys.path.insert(0, '/data/project/yixing-cutting-2026/question2')

import openpyxl
from openpyxl.utils import get_column_letter
from ems_solver_optimal import solve, print_solution, WORKPIECES, MATERIALS
import time

# 运行求解器
print("启动求解器...")
t0 = time.time()
sol = solve(num_trials=48, ils_iterations=800)
print_solution(sol)
print(f"\n总耗时: {time.time()-t0:.1f}s")

# 打开模板
template = '/data/project/yixing-cutting-2026/question2/result2.xlsx'
wb = openpyxl.load_workbook(template)
ws = wb['Sheet1']

# 构建块名称 → 工件计数的映射
# 模板行序: L01_1..L01_5, L02_1..L02_5, L03_1..L03_5
block_order = []
for m in MATERIALS:
    for i in range(1, m.quantity + 1):
        block_order.append(f"{m.name}_{i}")

# 填充数据 (row 4-18, col C-I for J01-J07)
for row_idx, bname in enumerate(block_order):
    excel_row = 4 + row_idx  # row 4-18
    pk_data = sol['results'].get(bname)
    if pk_data:
        for wp_idx, wp in enumerate(WORKPIECES):
            # Count occurrences of wp.name in this block
            cnt = sum(1 for name, *_ in pk_data['placements'] if name == wp.name)
            col = 3 + wp_idx  # column C=3, D=4, ..., I=9
            ws.cell(row=excel_row, column=col, value=cnt)

# 写入总结信息到 Sheet2
ws2 = wb['Sheet2']
ws2['A1'] = '子问题 2 求解结果 (v3 多策略两阶段 EMS + ILS)'
ws2['A3'] = '总利润'
ws2['B3'] = sol['total_profit']
ws2['A4'] = '材料利用率'
ws2['B4'] = f"{sol['utilization']:.4%}"
ws2['A5'] = '总工件数'
ws2['B5'] = sol['total_items']
ws2['A6'] = '已用体积'
ws2['B6'] = sol['total_used']
ws2['A7'] = '废料体积'
ws2['B7'] = sol['total_waste']
ws2['A8'] = '理论利润上界'
ws2['B8'] = sol['upper_bound']
ws2['A9'] = '占上界比例'
ws2['B9'] = f"{sol['ub_gap']:.4%}"
ws2['A10'] = '耗时 (秒)'
ws2['B10'] = f"{sol['elapsed']:.1f}"
ws2['A12'] = '工件产量明细'
ws2['A13'] = '工件'
ws2['B13'] = '数量'
ws2['C13'] = '满足 ≥10?'
for i, wp in enumerate(WORKPIECES):
    c = sol['counts'].get(wp.name, 0)
    ws2.cell(row=14+i, column=1, value=wp.name)
    ws2.cell(row=14+i, column=2, value=c)
    ws2.cell(row=14+i, column=3, value='是' if c >= 10 else '否')

# 保存
wb.save(template)
print(f"\n结果已导出到: {template}")
